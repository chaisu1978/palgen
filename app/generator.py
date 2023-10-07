"""
UI Palette based on HSB brand color
"""
import os
from PIL import Image, ImageDraw, ImageFont
import colorsys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Color
from openpyxl.utils import get_column_letter
import json

def get_all_palettes():
    palettes = []
    for root, dirs, files in os.walk('app/files'):
        for dir in dirs:
            palette = {}
            palette['name'] = dir
            palette['files'] = []
            for file in os.listdir(os.path.join(root, dir)):
                palette['files'].append(file)
            palettes.append(palette)
    return palettes

def get_palette_files(palette_name):
    files = []
    for root, dirs, files in os.walk('app/files'):
        for dir in dirs:
            if dir == palette_name:
                for file in os.listdir(os.path.join(root, dir)):
                    files.append(file)
    return files

def delete_palette(palette_name):
    for root, dirs, files in os.walk('app/files'):
        for dir in dirs:
            if dir == palette_name:
                for file in os.listdir(os.path.join(root, dir)):
                    os.remove(os.path.join(root, dir, file))
                os.rmdir(os.path.join(root, dir))

                return True
    return False

def hsb_to_rgb(hsb):
    rgb = tuple(int(x * 255) for x in colorsys.hsv_to_rgb(hsb[0] / 360, hsb[1] / 100, hsb[2] / 100))
    return rgb

def rgb_to_hsb(rgb):
    hsb = colorsys.rgb_to_hsv(rgb[0] / 255, rgb[1] / 255, rgb[2] / 255)
    return (hsb[0] * 360, hsb[1] * 100, hsb[2] * 100)

def hsb_to_hex(hsb):
    return '#%02x%02x%02x' % hsb_to_rgb(hsb)

def hex_to_rgb(hex):
    hex = hex.strip('#')
    hex = tuple(int(hex[i:i+2], 16) for i in (0, 2, 4))
    return hex

def hex_to_hsb(hex):
    rgb = hex_to_rgb(hex)
    hsb = rgb_to_hsb(rgb)
    hsb = tuple(int(round(x)) for x in hsb)
    return hsb

def generate_shades(base_color):
    h, s, b = base_color
    shades = {}
    shades["500"] = base_color
    shades["100"] = (h, max(0, s - 40), min(b + 50, 100))
    shades["900"] = (h, min(s + 40, 100), max(b - 70, 0))
    shades["300"] = (h, max(0, s - 20), min(b + 25, 100))
    shades["700"] = (h, min(s + 20, 100), max(b - 35, 0))
    shades["200"] = (h, max(0, s - 30), min(b + 37.5, 100))
    shades["400"] = (h, max(0, s - 10), min(b + 12.5, 100))
    shades["600"] = (h, min(s + 10, 100), max(b - 17.5, 0))
    shades["800"] = (h, min(s + 30, 100), max(b - 52.5, 0))
    shades = dict(sorted(shades.items(), key=lambda item: int(item[0])))
    return shades

def brightness(rgb):
    return (rgb[0]*299 + rgb[1]*587 + rgb[2]*2.2) / 1000

def contrast_color(rgb):
    return (0, 0, 0) if brightness(rgb) > 128 else (255, 255, 255)

def generate_palette(primary, secondary=None, tertiary=None):
    primary_hsb = hex_to_hsb(primary)
    secondary_hsb = hex_to_hsb(secondary) if secondary else None
    tertiary_hsb = hex_to_hsb(tertiary) if tertiary else None
    palette = {}
    palette["Primary"] = generate_shades(primary_hsb)
    if secondary_hsb:
        palette["Secondary"] = generate_shades(secondary_hsb)
    if tertiary_hsb:
        palette["Tertiary"] = generate_shades(tertiary_hsb)
    palette["Neutral"] = generate_shades((primary_hsb[0], 15, 70))

    supporting_colors = {
        "Green": (134, (primary_hsb[1] - 10), (primary_hsb[2] - 2)),
        "Orange": (23, (primary_hsb[1] - 10), (primary_hsb[2] - 2)),
        "Red": (0, (primary_hsb[1] - 10), (primary_hsb[2] - 2)),
        "Blue": (204, (primary_hsb[1] - 10), (primary_hsb[2] - 2))
    }

    for color in supporting_colors:
        palette[color] = generate_shades(supporting_colors[color])

    # Convert lists to tuples
    for name, values in palette.items():
        for shade, hsb in values.items():
            palette[name][shade] = tuple(hsb)

    return palette


class PaletteGenerator:
    def __init__(self, name, primary, secondary=None, tertiary=None):
        self.name = name
        self.primary = primary
        self.secondary = secondary
        self.tertiary = tertiary
        self.user_data = {}
        self.palette = generate_palette(self.primary, self.secondary, self.tertiary)
        self.create_directories()
        self.generate_files()
        self.user_data = {
            'name': self.name,
            'primary': self.primary,
            'secondary': self.secondary,
            'tertiary': self.tertiary,
            'files': [f'{self.name}-color-palette.png', f'{self.name}-color-pallete.xlsx', f'{self.name}-color-palette.css', f'{self.name}-color-palette.dart', f'{self.name}-color-palette.json'],
            'png_file': f'{self.name}-color-palette.png',
        }

    def create_directories(self):
        if not os.path.exists('app/files'):
            os.mkdir('app/files')
        palette_dir = os.path.join('app/files', self.name)
        if not os.path.exists(palette_dir):
            os.mkdir(palette_dir)

    def generate_files(self):
        self.generate_png_image()
        self.generate_excel_file()
        self.generate_css_file()
        self.generate_dart_file()
        self.generate_json_file()

    def generate_png_image(self):
        img_width = int(11.69 * 150)
        img_height = int(8.27 * 150)
        img = Image.new('RGB', (img_width, img_height), color='white')
        d = ImageDraw.Draw(img)
        font = ImageFont.truetype("app/arial.ttf", 15)
        total_colors = len(self.palette)

        def draw_rectangle(draw, rgb, text, name, i, j):
            x1 = j * (img_width // total_colors) + 2
            y1 = i * (img_height // len(self.palette[name])) + 2
            x2 = (j + 1) * (img_width // total_colors) - 2
            y2 = (i + 1) * (img_height // len(self.palette[name])) - 2
            draw.rectangle([(x1, y1), (x2, y2)], fill=rgb)

            text_width, text_height = draw.textbbox((0, 0), text, font=font)[2:]
            text_x = x1 + ((img_width // total_colors) - text_width) / 2
            text_y = y1 + ((img_height // len(self.palette[name])) - text_height) / 2

            text_color = contrast_color(rgb)

            draw.text((text_x, text_y), text, fill=text_color, font=font)

            if i == len(self.palette[name]) - 1:
                name_width, name_height = draw.textbbox((0, 0), name, font=font)[2:]
                name_x = x1 + ((img_width // total_colors) - name_width) / 2
                name_y = y2 + 5
                draw.text((name_x, name_y), name, fill=(0, 0, 0), font=font)

        for j, (name, palette) in enumerate(self.palette.items()):
            for i, shade in enumerate(palette):
                rgb = hsb_to_rgb(palette[shade])
                hex_rgb = hsb_to_hex(palette[shade])
                text = f"Weight - {shade}\nHSB - {palette[shade]}\nRGB - {rgb}\nHEX - {hex_rgb}"
                draw_rectangle(d, rgb, text, name, i, j)

        # Draw palette names in the last row
        for j, (name, palette) in enumerate(self.palette.items()):
            rgb = hsb_to_rgb(palette["500"])
            hex_rgb = hsb_to_hex(palette["500"])
            text = f"{name}"
            draw_rectangle(d, rgb, text, name, len(palette) - 1, j)

        img.save(f'app/files/{self.name}/{self.name}-color-palette.png')

    def generate_excel_file(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        total_colors = len(self.palette)
        shade_row_mapping = {shade: i + 2 for i, shade in enumerate(["100", "200", "300", "400", "500", "600", "700", "800", "900"])}

        for j, (name, palette) in enumerate(self.palette.items()):
            ws.column_dimensions[get_column_letter(j + 2)].width = 30
            ws.cell(row=1, column=j + 2, value=name)

            for shade, hsb in palette.items():
                ws.row_dimensions[shade_row_mapping[shade]].height = 100
                rgb = hsb_to_rgb(hsb)
                hex_rgb = hsb_to_hex(hsb)
                fill = PatternFill(start_color=hex_rgb[1:], end_color=hex_rgb[1:], fill_type="solid")
                cell = ws.cell(row=shade_row_mapping[shade], column=j + 2)
                cell.fill = fill
                font_color_rgb = contrast_color(rgb)
                font_color_rgb_str = f'{font_color_rgb[0]:02X}{font_color_rgb[1]:02X}{font_color_rgb[2]:02X}'
                font_color = Font(color=Color(rgb=font_color_rgb_str))
                cell.font = font_color
                alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.alignment = alignment
                font_color_rgb = contrast_color(rgb)
                font_color_rgb_str = f'{font_color_rgb[0]:02X}{font_color_rgb[1]:02X}{font_color_rgb[2]:02X}'
                font_color = Font(color=Color(rgb=font_color_rgb_str))
                cell.font = font_color
                cell.value = f"Weight - {shade}\nHSB - {hsb}\nRGB - {rgb}\nHEX - {hex_rgb}"

        wb.save(f'app/files/{self.name}/{self.name}-color-pallete.xlsx')

    def generate_css_file(self):
        css_filename = f'app/files/{self.name}/{self.name}-color-palette.css'
        with open(css_filename, 'w') as css_file:
            css_file.write('/* EXAMPLE USAGE */\n')
            css_file.write('/* .header { */\n')
            css_file.write('/* background-color: var(--primary-500); */\n')
            css_file.write('/* color: var(--neutral-900); */\n')
            css_file.write('/* } */\n\n')

            css_file.write('/* Button Styles */\n')
            css_file.write('/* .button { */\n')
            css_file.write('/*   background-color: var(--secondary-500); */\n')
            css_file.write('/*   color: var(--neutral-100); */\n')
            css_file.write('/* } */\n\n')

            css_file.write('/* Color Variables */\n')
            css_file.write(':root {\n')
            for name, palette in self.palette.items():
                for shade, hsb in palette.items():
                    hex_rgb = hsb_to_hex(hsb)
                    css_file.write(f'  --{name.lower()}-{shade}: {hex_rgb};\n')
            css_file.write('}\n')

    def generate_dart_file(self):
        dart_filename = f'app/files/{self.name}/{self.name}-color-palette.dart'

        with open(dart_filename, 'w') as dart_file:
            dart_file.write("import 'package:flutter/material.dart';\n\n")
            dart_file.write("class MyPalette {\n")

            # Define your color constants
            for name, palette in self.palette.items():
                for shade, hsb in palette.items():
                    hex_rgb = hsb_to_hex(hsb)
                    dart_file.write(
                        f"  static const Color {name.lower()}{shade} = Color(0x{hex_rgb[1:]});\n")

            dart_file.write("\n")
            dart_file.write("  // Define ThemeData\n")
            dart_file.write("  static ThemeData myAppTheme = ThemeData(\n")
            dart_file.write("    primaryColor: Primary500,  // Replace with your primary color\n")
            dart_file.write("    accentColor: Secondary500,  // Replace with your secondary color\n")
            dart_file.write("    // Define other theme properties as needed\n")
            dart_file.write("  );\n")

            dart_file.write("\n")
            dart_file.write("  // Define custom text styles\n")
            for name, palette in self.palette.items():
                for shade, hsb in palette.items():
                    dart_file.write(
                        f"  static TextStyle {name.lower()}{shade}TextStyle = TextStyle(\n")
                    dart_file.write(f"    color: {name.lower()}{shade},\n")
                    dart_file.write("    // Define text styling properties as needed\n")
                    dart_file.write("  );\n")

            dart_file.write("}\n")

    def generate_json_file(self):
        # Create the JSON data dictionary
        json_data = {
            'name': self.name,
            'palette_hsl': {},
        }

        # Add each color palette to the JSON data
        for palette_name, palette_values in self.palette.items():
            json_data['palette_hsl'][palette_name] = {
                level: hsl for level, hsl in palette_values.items()
            }

        # Define the output JSON filename
        json_filename = f'app/files/{self.name}/{self.name}-color-palette.json'

        # Write the JSON data to the file with proper indentation
        with open(json_filename, 'w') as json_file:
            json.dump(json_data, json_file, indent=4)
