"""
UI Palette based on HSB brand color
"""
from django.conf import settings
import os
import shutil
from pathlib import Path
from PIL import Image, ImageDraw, ImageFont
import colorsys
import openpyxl
from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment,
    Color
    )
from openpyxl.utils import get_column_letter
import json


def get_all_palettes():
    palettes = []
    for root, dirs, files in os.walk('app/files'):
        for dir in dirs:
            palette = {}
            palette['name'] = dir
            palette['files'] = []
            for file in os.listdir(os.path.join(root, dir)):
                palette['files'].append(file)
            palettes.append(palette)
    return palettes


def hsb_to_rgb(hsb):
    rgb = tuple(int(x * 255) for x in colorsys.hsv_to_rgb(hsb[0] / 360, hsb[1] / 100, hsb[2] / 100))  # noqa: E501
    return rgb


def rgb_to_hsb(rgb):
    hsb = colorsys.rgb_to_hsv(rgb[0] / 255, rgb[1] / 255, rgb[2] / 255)
    return (hsb[0] * 360, hsb[1] * 100, hsb[2] * 100)


def hsb_to_hex(hsb):
    return '#%02x%02x%02x' % hsb_to_rgb(hsb)


def hex_to_rgb(hex):
    hex = hex.strip('#')
    hex = tuple(int(hex[i:i+2], 16) for i in (0, 2, 4))
    return hex


def hex_to_hsb(hex):
    rgb = hex_to_rgb(hex)
    hsb = rgb_to_hsb(rgb)
    hsb = tuple(int(round(x)) for x in hsb)
    return hsb


def generate_shades(base_color):
    h, s, b = base_color
    shades = {}
    shades["500"] = base_color
    shades["100"] = (h, max(0, s - 40), min(b + 50, 100))
    shades["900"] = (h, min(s + 40, 100), max(b - 70, 0))
    shades["300"] = (h, max(0, s - 20), min(b + 25, 100))
    shades["700"] = (h, min(s + 20, 100), max(b - 35, 0))
    shades["200"] = (h, max(0, s - 30), min(b + 37.5, 100))
    shades["400"] = (h, max(0, s - 10), min(b + 12.5, 100))
    shades["600"] = (h, min(s + 10, 100), max(b - 17.5, 0))
    shades["800"] = (h, min(s + 30, 100), max(b - 52.5, 0))
    shades = dict(sorted(shades.items(), key=lambda item: int(item[0])))
    return shades


def brightness(rgb):
    return (rgb[0]*299 + rgb[1]*587 + rgb[2]*2.2) / 1000


def contrast_color(rgb):
    return (0, 0, 0) if brightness(rgb) > 128 else (255, 255, 255)


def rgb_to_cmyk(rgb):
    """
    Convert RGB values (0-255) to CMYK values (0-100)
    """
    r, g, b = [x / 255.0 for x in rgb]
    
    # Calculate K (black)
    k = 1 - max(r, g, b)
    
    # Handle the case where k = 1 (pure black)
    if k == 1:
        return (0, 0, 0, 100)
    
    # Calculate C, M, Y
    c = (1 - r - k) / (1 - k)
    m = (1 - g - k) / (1 - k)
    y = (1 - b - k) / (1 - k)
    
    # Convert to percentages and round
    return (
        int(round(c * 100)),
        int(round(m * 100)),
        int(round(y * 100)),
        int(round(k * 100))
    )


def generate_palette(primary, secondary=None, tertiary=None):
    primary_hsb = hex_to_hsb(primary)
    secondary_hsb = hex_to_hsb(secondary) if secondary else None
    tertiary_hsb = hex_to_hsb(tertiary) if tertiary else None
    palette = {}
    palette["Primary"] = generate_shades(primary_hsb)
    if secondary_hsb:
        palette["Secondary"] = generate_shades(secondary_hsb)
    if tertiary_hsb:
        palette["Tertiary"] = generate_shades(tertiary_hsb)
    palette["Neutral"] = generate_shades((primary_hsb[0], 15, 70))

    supporting_colors = {
        "Green": (134, (primary_hsb[1] - 10), (primary_hsb[2] - 2)),
        "Orange": (23, (primary_hsb[1] - 10), (primary_hsb[2] - 2)),
        "Red": (0, (primary_hsb[1] - 10), (primary_hsb[2] - 2)),
        "Blue": (204, (primary_hsb[1] - 10), (primary_hsb[2] - 2))
    }

    for color in supporting_colors:
        palette[color] = generate_shades(supporting_colors[color])

    # Convert lists to tuples
    for name, values in palette.items():
        for shade, hsb in values.items():
            palette[name][shade] = tuple(hsb)

    return palette


class PaletteGenerator:
    def __init__(self, name, primary, secondary=None, tertiary=None, base_dir=None):
        self.name = name
        self.primary = primary
        self.secondary = secondary
        self.tertiary = tertiary
        self.base_dir = base_dir or os.path.join(settings.MEDIA_ROOT, 'palettes')
        self.output_dir = self.base_dir  # Use the provided base_dir directly
        self.user_data = {}

        # Ensure output directory exists
        os.makedirs(self.output_dir, exist_ok=True)

        # Generate the color palette
        self.palette = generate_palette(
            self.primary,
            self.secondary,
            self.tertiary
        )

        # Generate all files
        self.generate_files()

        # Set user data with file paths relative to MEDIA_ROOT
        self.user_data = {
            'name': self.name,
            'primary': self.primary,
            'secondary': self.secondary,
            'tertiary': self.tertiary,
            'output_dir': self.output_dir,
            'files': self.get_generated_files()
        }

    def get_output_path(self, extension):
        """Get the full output path for a file with the given extension"""
        filename = f'{self.name}-color-palette.{extension}'
        return os.path.join(self.output_dir, filename)

    def get_generated_files(self):
        """Return a list of all generated files with their relative paths"""
        file_extensions = ['png', 'xlsx', 'css', 'ts', 'dart']
        files = []

        for ext in file_extensions:
            file_path = self.get_output_path(ext)
            if os.path.exists(file_path):
                rel_path = os.path.relpath(file_path, settings.MEDIA_ROOT)
                files.append({
                    'name': os.path.basename(file_path),
                    'path': rel_path,
                    'url': os.path.join(settings.MEDIA_URL, rel_path).replace('\\', '/')
                })

        return files

    def generate_files(self):
        self.generate_png_image()
        self.generate_excel_file()
        self.generate_css_file()
        self.generate_typescript_file()
        self.generate_dart_file()

    def generate_png_image(self):
        img_width = int(11.69 * 150)
        img_height = int(8.27 * 150)
        img = Image.new('RGB', (img_width, img_height), color='white')
        d = ImageDraw.Draw(img)

        # Use the arial.ttf from the generator directory
        font_path = os.path.join(os.path.dirname(__file__), 'arial.ttf')
        font = ImageFont.truetype(font_path, 15)
        font_bold = ImageFont.truetype(font_path, 16)  # Slightly larger for bold effect
        total_colors = len(self.palette)

        def draw_rectangle(draw, rgb, cmyk, text_lines, cmyk_text, name, i, j):
            x1 = j * (img_width // total_colors) + 2
            y1 = i * (img_height // len(self.palette[name])) + 2
            x2 = (j + 1) * (img_width // total_colors) - 2
            y2 = (i + 1) * (img_height // len(self.palette[name])) - 2
            draw.rectangle([(x1, y1), (x2, y2)], fill=rgb)

            text_color = contrast_color(rgb)
            
            # Calculate line height
            line_height = 18
            total_height = len(text_lines) * line_height + line_height  # +1 for bold CMYK line
            start_y = y1 + ((y2 - y1) - total_height) / 2
            
            # Draw regular text lines
            for idx, line in enumerate(text_lines):
                line_bbox = draw.textbbox((0, 0), line, font=font)
                line_width = line_bbox[2] - line_bbox[0]
                text_x = x1 + ((img_width // total_colors) - line_width) / 2
                text_y = start_y + (idx * line_height)
                draw.text((text_x, text_y), line, fill=text_color, font=font)
            
            # Draw bold CMYK line
            cmyk_bbox = draw.textbbox((0, 0), cmyk_text, font=font_bold)
            cmyk_width = cmyk_bbox[2] - cmyk_bbox[0]
            cmyk_x = x1 + ((img_width // total_colors) - cmyk_width) / 2
            cmyk_y = start_y + (len(text_lines) * line_height)
            # Draw text twice with slight offset for bold effect
            draw.text((cmyk_x, cmyk_y), cmyk_text, fill=text_color, font=font_bold)
            draw.text((cmyk_x + 0.5, cmyk_y), cmyk_text, fill=text_color, font=font_bold)

            if i == len(self.palette[name]) - 1:
                name_width, name_height = draw.textbbox(
                    (0, 0),
                    name,
                    font=font
                    )[2:]
                name_x = x1 + ((img_width // total_colors) - name_width) / 2
                name_y = y2 + 5
                draw.text((name_x, name_y), name, fill=(0, 0, 0), font=font)

        for j, (name, palette) in enumerate(self.palette.items()):
            for i, shade in enumerate(palette):
                rgb = hsb_to_rgb(palette[shade])
                hex_rgb = hsb_to_hex(palette[shade])
                cmyk = rgb_to_cmyk(rgb)
                
                # Prepare text lines (non-bold)
                text_lines = [
                    f"Weight - {shade}",
                    f"HSB - {palette[shade]}",
                    f"RGB - {rgb}",
                    f"HEX - {hex_rgb}"
                ]
                
                # CMYK line (will be bold)
                cmyk_text = f"CMYK - {cmyk}"
                
                draw_rectangle(d, rgb, cmyk, text_lines, cmyk_text, name, i, j)

        output_path = self.get_output_path('png')
        img.save(output_path)

    def generate_excel_file(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        shade_row_mapping = {shade: i + 2 for i, shade in enumerate([
            "100",
            "200",
            "300",
            "400",
            "500",
            "600",
            "700",
            "800",
            "900"
            ])}

        for j, (name, palette) in enumerate(self.palette.items()):
            ws.column_dimensions[get_column_letter(j + 2)].width = 30
            ws.cell(row=1, column=j + 2, value=name)

            for shade, hsb in palette.items():
                ws.row_dimensions[shade_row_mapping[shade]].height = 100
                rgb = hsb_to_rgb(hsb)
                hex_rgb = hsb_to_hex(hsb)
                cmyk = rgb_to_cmyk(rgb)
                fill = PatternFill(start_color=hex_rgb[1:], end_color=hex_rgb[1:], fill_type="solid")  # noqa: E501
                cell = ws.cell(row=shade_row_mapping[shade], column=j + 2)
                cell.fill = fill
                font_color_rgb = contrast_color(rgb)
                font_color_rgb_str = f'{font_color_rgb[0]:02X}{font_color_rgb[1]:02X}{font_color_rgb[2]:02X}'  # noqa: E501
                font_color = Font(color=Color(rgb=font_color_rgb_str))
                cell.font = font_color
                alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # noqa: E501
                cell.alignment = alignment
                cell.value = f"Weight - {shade}\nHSB - {hsb}\nRGB - {rgb}\nHEX - {hex_rgb}\nCMYK - {cmyk}"  # noqa: E501

        output_path = self.get_output_path('xlsx')
        wb.save(output_path)

    def generate_css_file(self):
        output_path = self.get_output_path('css')
        with open(output_path, 'w') as f:
            f.write('/* EXAMPLE USAGE */\n')
            f.write('/* .header { */\n')
            f.write('/* background-color: var(--primary-500); */\n')
            f.write('/* color: var(--neutral-900); */\n')
            f.write('/* } */\n\n')

            f.write('/* Button Styles */\n')
            f.write('/* .button { */\n')
            f.write('/*   background-color: var(--secondary-500); */\n')
            f.write('/*   color: var(--neutral-100); */\n')
            f.write('/* } */\n\n')

            f.write('/* Color Variables */\n')
            f.write(':root {\n')
            for name, palette in self.palette.items():
                for shade, hsb in palette.items():
                    hex_rgb = hsb_to_hex(hsb)
                    f.write(f'  --{name.lower()}-{shade}: {hex_rgb};\n')
            f.write('}\n')

    def generate_dart_file(self):
        """
        Generate a Dart file compatible with Flutter Material theming
        Similar structure to TypeScript paletteComponents
        """
        output_path = self.get_output_path('dart')
        with open(output_path, 'w') as f:
            f.write("import 'package:flutter/material.dart';\n\n")
            f.write("// Centralized shared components for theme configuration\n")
            f.write("class PaletteComponents {\n")
            
            # Generate color maps for each palette
            for name, palette in self.palette.items():
                f.write(f"  static const Map<int, Color> {name.lower()} = {{\n")
                
                for shade, hsb in palette.items():
                    hex_rgb = hsb_to_hex(hsb)
                    # Flutter Color format requires 0xFF prefix for full opacity
                    f.write(f"    {shade}: Color(0xFF{hex_rgb[1:].upper()}),\n")
                
                f.write("  };\n\n")
            
            # Generate MaterialColor swatches for primary colors
            f.write("  // Material Color Swatches\n")
            for name in ['primary', 'secondary', 'tertiary']:
                if name.capitalize() in self.palette:
                    f.write(f"  static const MaterialColor {name}Swatch = MaterialColor(\n")
                    f.write(f"    0xFF{hsb_to_hex(self.palette[name.capitalize()]['500'])[1:].upper()},\n")
                    f.write(f"    {name},\n")
                    f.write("  );\n\n")
            
            f.write("}\n")

    def generate_typescript_file(self):
        """
        Generate a TypeScript file compatible with MUI theming
        Similar to frontend/src/themes/paletteComponents.ts
        """
        output_path = self.get_output_path('ts')
        with open(output_path, 'w') as f:
            f.write('// Centralized shared components for theme configuration\n')
            f.write('const paletteComponents = {\n')
            
            # Generate each color palette
            for idx, (name, palette) in enumerate(self.palette.items()):
                f.write(f'  {name.lower()}: {{\n')
                
                # Write each shade
                for shade, hsb in palette.items():
                    hex_rgb = hsb_to_hex(hsb)
                    f.write(f"    {shade}: '{hex_rgb}',\n")
                
                # Close the color object
                if idx < len(self.palette) - 1:
                    f.write('  },\n')
                else:
                    f.write('  },\n')
            
            f.write('};\n\n')
            f.write('export default paletteComponents;\n')
