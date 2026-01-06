# In generator.py

from django.conf import settings
import os
import shutil
from pathlib import Path
from PIL import Image, ImageDraw, ImageFont
import colorsys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Color
from openpyxl.utils import get_column_letter
import json

class PaletteGenerator:
    def __init__(self, name, primary, secondary=None, tertiary=None, base_dir=None):
        self.name = name
        self.primary = primary
        self.secondary = secondary
        self.tertiary = tertiary
        self.base_dir = base_dir or os.path.join(settings.MEDIA_ROOT, 'palettes')
        self.output_dir = os.path.join(self.base_dir, name)
        self.user_data = {}

        # Ensure output directory exists
        os.makedirs(self.output_dir, exist_ok=True)

        # Generate the color palette
        self.palette = generate_palette(
            self.primary,
            self.secondary,
            self.tertiary
        )

        # Generate all files
        self.generate_files()

        # Set user data with file paths relative to MEDIA_ROOT
        self.user_data = {
            'name': self.name,
            'primary': self.primary,
            'secondary': self.secondary,
            'tertiary': self.tertiary,
            'output_dir': self.output_dir,
            'files': self.get_generated_files()
        }

    def get_generated_files(self):
        """Return a list of all generated files with their relative paths"""
        file_extensions = ['png', 'xlsx', 'css', 'dart', 'json']
        files = []

        for ext in file_extensions:
            filename = f'{self.name}-color-palette.{ext}'
            filepath = os.path.join(self.output_dir, filename)
            if os.path.exists(filepath):
                rel_path = os.path.relpath(filepath, settings.MEDIA_ROOT)
                files.append({
                    'name': filename,
                    'path': rel_path,
                    'url': os.path.join(settings.MEDIA_URL, rel_path).replace('\\', '/')
                })

        return files

    def generate_files(self):
        """Generate all palette files"""
        self.generate_png_image()
        self.generate_excel_file()
        self.generate_css_file()
        self.generate_dart_file()
        self.generate_json_file()

    def get_output_path(self, extension):
        """Get the full output path for a file with the given extension"""
        filename = f'{self.name}-color-palette.{extension}'
        return os.path.join(self.output_dir, filename)

    def generate_png_image(self):
        img_width = int(11.69 * 150)
        img_height = int(8.27 * 150)
        img = Image.new('RGB', (img_width, img_height), color='white')
        d = ImageDraw.Draw(img)

        # Use the arial.ttf from the generator directory
        font_path = os.path.join(os.path.dirname(__file__), 'arial.ttf')
        font = ImageFont.truetype(font_path, 15)

        total_colors = len(self.palette)

        def draw_rectangle(draw, rgb, text, name, i, j):
            x1 = j * (img_width // total_colors) + 2
            y1 = i * (img_height // len(self.palette[name])) + 2
            x2 = (j + 1) * (img_width // total_colors) - 2
            y2 = (i + 1) * (img_height // len(self.palette[name])) - 2
            draw.rectangle([(x1, y1), (x2, y2)], fill=rgb)

            text_width, text_height = draw.textbbox((0, 0), text, font=font)[2:]
            text_x = x1 + ((img_width // total_colors) - text_width) / 2
            text_y = y1 + ((img_height // len(self.palette[name])) - text_height) / 2

            text_color = contrast_color(rgb)
            draw.text((text_x, text_y), text, fill=text_color, font=font)

        # Rest of the PNG generation code...
        # Save the image
        output_path = self.get_output_path('png')
        img.save(output_path)

    def generate_excel_file(self):
        output_path = self.get_output_path('xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Color Palette"

        # ... rest of the Excel generation code ...

        wb.save(output_path)

    def generate_css_file(self):
        output_path = self.get_output_path('css')
        with open(output_path, 'w') as f:
            # ... CSS generation code ...
            pass

    def generate_dart_file(self):
        output_path = self.get_output_path('dart')
        with open(output_path, 'w') as f:
            # ... Dart generation code ...
            pass

    def generate_json_file(self):
        output_path = self.get_output_path('json')
        with open(output_path, 'w') as f:
            json.dump(self.palette, f, indent=2)

    def cleanup(self):
        """Clean up generated files"""
        if os.path.exists(self.output_dir):
            shutil.rmtree(self.output_dir)





# In your view
generator = PaletteGenerator(
    name=palette.name,
    primary=palette.primary,
    secondary=palette.secondary,
    tertiary=palette.tertiary,
    base_dir=os.path.join(settings.MEDIA_ROOT, 'palettes', str(palette.id))
)

# Get the list of generated files
generated_files = generator.get_generated_files()
